import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.chart import Reference, BarChart3D, PieChart3D
from openpyxl.styles import Font, PatternFill, Alignment
import signal
import sys


def handler_signal(signal, frame):
    print("\n\n Interrupción! Saliendo del programa de manera ordenada")
    sys.exit(1)

# Señal de control por si el usuario introduce Ctrl + C para parar el programa
signal.signal(signal.SIGINT, handler_signal)


def procesar_pedidos(pedidos):

    pedidos = pedidos[pedidos['pizza_id'].isnull() == False]
    pedidos = pedidos[pedidos['quantity'].isnull() == False]
    pizza_id = []
    quantity = []

    # Ahora procesaremos estas dos columnas del dataframe, remplazando ciertos caracteres por otros
    # para así tener estas dos columnas de manera que se correspondan los nombres de las pizzas
    # con los nombres de las pizzas en los dataframe pizzas e ingredientes. Con ello tratamos de
    # tener los nombres de las pizzas en el formato <nombre_separado_por_guiones_bajos_tamaño_pizza>
    # y las cantidades en números enteros positivos.

    for pizza_sin_procesar in pedidos['pizza_id']:
        pizza = pizza_sin_procesar.replace('@', 'a').replace('3', 'e').replace('0', 'o').replace('-', '_').replace(' ', '_')
        pizza_id.append(pizza)

    for cantidad in pedidos['quantity']:    
        cantidad = cantidad.replace('One', '1').replace('one', '1').replace('two', '2')
        cantidad = abs(int(cantidad))
        quantity.append(cantidad)

    pedidos['pizza_id'] = pizza_id  # Reescribimos las dos columnas del dataframe que acabamos de procesar
    pedidos['quantity'] = quantity

    return pedidos


def añadir_pizzas_totales(pizzas, pedidos):

    # Al dataframe de pizzas añade una columna que cuente las veces que se ha pedido cada pizza en el año

    lista = []
    for pizza in pizzas['pizza_id']:
        lista.append(pedidos[pedidos['pizza_id'] == pizza]['quantity'].sum())
    pizzas['quantity sold'] = lista

    return pizzas


def cargar_ficheros():

    # Carga y procesa los ficheros csv necesarios para generar las gráfias a modo de análisis

    compra_semanal_ingredientes = pd.read_csv("compra_semanal_ingredientes.csv", sep = ",", encoding = "UTF-8")
    pedidos = pd.read_csv("order_details.csv", sep = ";", encoding = "UTF-8")
    pedidos = procesar_pedidos(pedidos)
    pizzas = pd.read_csv("pizzas.csv", sep = ",", encoding = "UTF-8")
    pizzas = añadir_pizzas_totales(pizzas, pedidos)

    return compra_semanal_ingredientes, pizzas


def crear_writer(archivo_salida):

    # Esta función crea el fichero excel y crea un writer para poder droppear dataframes en él de manera cómoda

    wb = Workbook()
    wb.save(archivo_salida)
    wb = load_workbook(archivo_salida)
    writer = pd.ExcelWriter(archivo_salida, engine = 'openpyxl')
    writer.book = wb
    writer.sheets = {ws.title:ws for ws in wb.worksheets}

    return writer, wb  # Devuelve el writer y el workbook u objeto de trabajo


def colorear_mayores(min_fila, max_fila, pestaña, columna_principal, otras_columnas, leyenda, pos_leyenda):

    # Colorea las líneas que contienen la mayor cantidad de un dato en concreto (precio, pedidos...)
    # Como ordenar por una columna de un dataframe

    verde = PatternFill(patternType='solid', fgColor='35FC03')
    pestaña[pos_leyenda] = leyenda      # Creamos una leyenda para que se entienda lo que significa que una celda esté coloreada
    pestaña[pos_leyenda].fill = verde
    lista_mayores = [min_fila+1, min_fila+2, min_fila+3, min_fila+4, min_fila+5]

    for i in range (12, max_fila + 1):  # Primero saca las filas que habrá que colorear
        for j in range(len(lista_mayores)):
            pestaña[f'{columna_principal}{i}'].value
            pestaña[f'{columna_principal}{lista_mayores[j]}'].value
            if pestaña[f'{columna_principal}{i}'].value > pestaña[f'{columna_principal}{lista_mayores[j]}'].value:
                lista_mayores[j] = i
                break

    otras_columnas.append(columna_principal)

    for i in lista_mayores:             # Luego las colorea
        for columna in otras_columnas:
            pestaña[f'{columna}{str(i)}'].fill = verde


def colorear_menores(min_fila, max_fila, pestaña, columna_principal, otras_columnas, leyenda, pos_leyenda):

    # Igual que la función colorear_mayores pero con los menores valores.

    rojo = PatternFill(patternType='solid', fgColor='FC2C03')
    pestaña[pos_leyenda] = leyenda
    pestaña[pos_leyenda].fill = rojo
    lista_menores = [min_fila+1, min_fila+2, min_fila+3, min_fila+4, min_fila+5]

    for i in range (12, max_fila + 1):
        for j in range(len(lista_menores)):
            if pestaña[f'{columna_principal}{i}'].value < pestaña[f'{columna_principal}{lista_menores[j]}'].value and pestaña[f'{columna_principal}{i}'].value != 0:
                lista_menores[j] = i
                break

    otras_columnas.append(columna_principal)

    for i in lista_menores:
        for columna in otras_columnas:
            pestaña[f'{columna}{str(i)}'].fill = rojo


def columnas(wb, pestaña, tamaños):

    # Trabaja y organiza las filas y columnas del excel

    min_columna = wb.active.min_column
    max_columna = wb.active.max_column
    min_fila = wb.active.min_row
    max_fila = wb.active.max_row
    for key in tamaños.keys():
        pestaña.column_dimensions[key].width = tamaños[key]  # Hace alguna columna más grande según el dato que contenga, para que sea más visual

    return min_columna, max_columna, min_fila, max_fila  # Devuelve las dimensiones que ocupan los datos (el/los dataframes que se hallan cargado)
        

def grafico(chart, pestaña, min_columna, max_columna, min_fila, max_fila, titulo, estilo, ancho, altura, posicion):

    # Recibe un estilo de gráfico, la posicón de los datos que se quieren graficar y varias características del gráfico

    data = Reference(pestaña, min_col=min_columna+1, max_col = max_columna, min_row = min_fila, max_row = max_fila)
    categorias = Reference(pestaña, min_col=min_columna, max_col = min_columna, min_row = min_fila+1, max_row = max_fila)
    chart.add_data(data, titles_from_data = True)
    chart.set_categories(categorias)
    chart.title = titulo
    chart.style = estilo
    chart.width = ancho
    chart.height = altura
    pestaña.add_chart(chart, posicion)


def titular_pestaña (pestaña, merge):

    # Sirve para dar título a una pestaña del excel, y en general para hacer el excel más visual, ordenando y centrando los valores

    if merge:
        pestaña.merge_cells('B2:C2')
        pestaña.merge_cells('B3:C3')
    pestaña['B2'] = 'MAVEN PIZZAS'
    pestaña['B3'] = pestaña.title
    pestaña['B4'] = '2016'
    pestaña['B2'].font = Font('Times', bold = True, size = 17)
    pestaña['B3'].font = Font('Times', bold = True, size = 15)
    pestaña['B4'].font = Font('Times', bold = True, size = 12)
    for row in pestaña.iter_rows(min_row = 7, max_row = 120):
        for cell in row:
            cell.alignment = Alignment(horizontal='left') 


def pestaña_ingredientes(writer, wb, compra_semanal_ingredientes):

    # Emplea las funciones anteriores para crear una pestaña centrada en los ingredientes de las pizzas.
    
    compra_semanal_ingredientes.to_excel(writer, startrow = 5, startcol = 1, sheet_name = 'Ingredients Report', index = False)

    pestaña = wb['Ingredients Report']  # Carga la pestaña
    wb.active = pestaña
    pestaña.sheet_view.zoomScale = 60
    min_columna, max_columna, min_fila, max_fila = columnas(wb, pestaña, {'B': 25, 'C': 15, 'E': 25, 'F': 15})
    colorear_mayores(min_fila, max_fila, pestaña, 'C', ['B'], 'Most Used', 'C3')  # Colorea lo necesario, mayores y menores valores
    colorear_menores(min_fila, max_fila, pestaña, 'C', ['B'], 'Least Used', 'C4')
    grafico(BarChart3D(), pestaña, min_columna, max_columna, min_fila, max_fila, 'Ingredients', 6, 40, 12, 'E6') # Gráfico grande de barras
    titular_pestaña(pestaña, False)

    ingredientes_mas_usados = compra_semanal_ingredientes.sort_values('Amount (units)', ascending=False).head(5) # Filtramos el dataframe (top 5)
    ingredientes_mas_usados.to_excel(writer, startrow = 30, startcol = 4, sheet_name = 'Ingredients Report', index = False)
    grafico(PieChart3D(), pestaña, min_columna+3, max_columna+3, 31, 36, 'Top 5 Ingredients', 18, 14, 8, 'H31')  # Gráfico "Pie" con el top 5
    grafico(BarChart3D(), pestaña, min_columna+3, max_columna+3, 31, 36, 'Top 5 Ingredients', 21, 14, 8, 'Q31')  # Gráfico de barras con el top 5

    ingredientes_menos_usados = compra_semanal_ingredientes.sort_values('Amount (units)', ascending=True).head(5) # Filtramos el dataframe (bottom 5)
    ingredientes_menos_usados.to_excel(writer, startrow = 47, startcol = 4, sheet_name = 'Ingredients Report', index = False)
    grafico(PieChart3D(), pestaña, min_columna+3, max_columna+3, 48, 53, 'Bottom 5 Ingredients', 18, 14, 8, 'H48') # Gráfico "Pie" con el bottom 5
    grafico(BarChart3D(), pestaña, min_columna+3, max_columna+3, 48, 53, 'Bottom 5 Ingredients', 20, 14, 8, 'Q48') # Gráfico de barras con el bottom 5


def pestaña_pizzas_precios(writer, wb, pizzas):

    # Emplea las funciones anteriores para crear una pestaña centrada en los precios de las pizzas.
    # El funcionamiento es idéntico al de pestaña_ingredientes, adaptando los tamaños al dataframe de pizzas.

    pizzas = pizzas[['pizza_type_id', 'pizza_id', 'price', 'size', 'quantity sold']]
    pizzas.to_excel(writer, startrow = 5, startcol = 1, sheet_name = 'Pizzas Cost Report', index = False)

    pestaña2 = wb['Pizzas Cost Report']
    wb.active = pestaña2
    pestaña2.sheet_view.zoomScale = 60
    min_columna, max_columna, min_fila, max_fila = columnas(wb, pestaña2, {'B': 15, 'C': 15, 'E':6, 'F': 13, 'H': 15, 'I': 15, 'K':6, 'L': 13})
    colorear_mayores(min_fila, max_fila, pestaña2, 'D', ['B', 'C', 'E', 'F'], 'Expensive', 'D3')
    colorear_menores(min_fila, max_fila, pestaña2, 'D', ['B', 'C', 'E', 'F'], 'Cheap', 'D4')
    grafico(BarChart3D(), pestaña2, min_columna+1, max_columna-2, min_fila, max_fila, 'Pizza Price', 3, 60, 12, 'H6')
    titular_pestaña(pestaña2, True)

    pizzas_mas_caras = pizzas.sort_values('price', ascending=False).head(5)
    pizzas_mas_caras.to_excel(writer, startrow = 30, startcol = 7, sheet_name = 'Pizzas Cost Report', index = False)
    grafico(PieChart3D(), pestaña2, min_columna+7, max_columna+4, 31, 36, 'Top 5 Expensive Pizzas', 18, 14, 8, 'N31')
    grafico(BarChart3D(), pestaña2, min_columna+7, max_columna+4, 31, 36, 'Top 5 Expensive Pizzas', 21, 14, 8, 'W31')

    pizzas_menos_caras = pizzas.sort_values('price', ascending=True).head(5)
    pizzas_menos_caras.to_excel(writer, startrow = 47, startcol = 7, sheet_name = 'Pizzas Cost Report', index = False)
    grafico(PieChart3D(), pestaña2, min_columna+7, max_columna+4, 48, 53, 'Top 5 Cheap Pizzas', 18, 14, 8, 'N48')
    grafico(BarChart3D(), pestaña2, min_columna+7, max_columna+4, 48, 53, 'Top 5 Cheap Pizzas', 20, 14, 8, 'W48')


def pestaña_pizzas_cantidad(writer, wb, pizzas):

    # Emplea las funciones anteriores para crear una pestaña centrada en la cantidad vendida de pizzas.
    # El funcionamiento es idéntico al de pestaña_ingredientes, adaptando los tamaños al dataframe de pizzas.

    pizzas = pizzas[['pizza_type_id', 'pizza_id', 'quantity sold', 'size', 'price']]
    pizzas.to_excel(writer, startrow = 5, startcol = 1, sheet_name = 'Pizzas Quantity Report', index = False)

    pestaña3 = wb['Pizzas Quantity Report']
    wb.active = pestaña3
    pestaña3.sheet_view.zoomScale = 60
    min_columna, max_columna, min_fila, max_fila = columnas(wb, pestaña3, {'B': 15, 'C': 15, 'D': 13, 'E':6, 'H': 15, 'I': 15, 'J': 13, 'K':6})
    colorear_mayores(min_fila, max_fila, pestaña3, 'D', ['B', 'C', 'F', 'E'], 'Most Sold', 'D3')
    colorear_menores(min_fila, max_fila, pestaña3, 'D', ['B', 'C', 'F', 'E'], 'Least Sold', 'D4')
    grafico(BarChart3D(), pestaña3, min_columna+1, max_columna-2, min_fila, max_fila, 'Pizza Quantity Sold', 7, 60, 12, 'H6')
    titular_pestaña(pestaña3, True)

    pizzas_mas_vendidas = pizzas.sort_values('quantity sold', ascending=False).head(5)
    pizzas_mas_vendidas.to_excel(writer, startrow = 30, startcol = 7, sheet_name = 'Pizzas Quantity Report', index = False)
    grafico(PieChart3D(), pestaña3, min_columna+7, max_columna+4, 31, 36, 'Top 5 Sold Pizzas', 18, 14, 8, 'N31')
    grafico(BarChart3D(), pestaña3, min_columna+7, max_columna+4, 31, 36, 'Top 5 Sold Pizzas', 21, 14, 8, 'W31')

    # Consideramos el bottom 5 como las menos vendidas, pero tienen que haber sido vendidas al menos una vez.
    # Por eso filtramos el dataframe cogiendo las 10 menos vendidas y luego quedándonos con las 5 mayores de esas 10, 
    # las otras 5 no se han llegado a vender.

    pizzas_menos_vendidas = pizzas.sort_values('quantity sold', ascending=True).head(10).tail(5)
    pizzas_menos_vendidas.to_excel(writer, startrow = 47, startcol = 7, sheet_name = 'Pizzas Quantity Report', index = False)
    grafico(PieChart3D(), pestaña3, min_columna+7, max_columna+4, 48, 53, 'Bottom 5 Sold Pizzas', 18, 14, 8, 'N48')
    grafico(BarChart3D(), pestaña3, min_columna+7, max_columna+4, 48, 53, 'Bottom 5 Sold Pizzas', 20, 14, 8, 'W48')


def automatizar_excel(compra_semanal_ingredientes, pizzas, archivo_salida):

    # Crea el writer y las 3 pestañas del excel, elimina una pestaña que se ha generado automáticamente y guarda el archivo

    writer, wb = crear_writer(archivo_salida)
    pestaña_ingredientes(writer, wb, compra_semanal_ingredientes)
    pestaña_pizzas_precios(writer, wb, pizzas)
    pestaña_pizzas_cantidad(writer, wb, pizzas)
    del wb['Sheet']
    wb.active = wb['Ingredients Report'] # Establece que la pestaña de ingredientes sea la que aparezca al abrir el archivo
    wb.save(archivo_salida)


if __name__ == '__main__':

    # Cargar los ficheros y crear el excel

    compra_semanal_ingredientes, pizzas = cargar_ficheros()
    automatizar_excel(compra_semanal_ingredientes, pizzas, 'Pizza_Reports.xlsx')